"""
Модуль анализа страниц с чертежами с использованием модели gemma4:31b через Ollama.
Анализирует каждую страницу и сохраняет результаты в отдельной папке сессии.
СОХРАНЕНИЕ В EXCEL, JSON, TXT, CSV С ОБЯЗАТЕЛЬНЫМ УКАЗАНИЕМ ЛИНЕЙНЫХ РАЗМЕРОВ И МАТЕРИАЛОВ
"""

import os
import ollama
import json
import base64
import time
import logging
from pathlib import Path
from typing import List, Dict, Any
import fitz  # PyMuPDF для работы с PDF
from PIL import Image, ImageEnhance, ImageFilter
import io
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================
# НАСТРОЙКА ЛОГИРОВАНИЯ
# ============================================

def setup_logging():
    current_dir = Path(__file__).resolve().parent  # scripts/
    project_dir = current_dir.parent  # корень проекта (Draw_Inspect)
    log_dir = project_dir / 'outputs' / 'logs'
    
    log_dir.mkdir(parents=True, exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_dir / 'page_analyzer.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# Конфигурация Ollama
OLLAMA_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
ANALYSIS_MODEL = os.getenv("DRAWING_ANALYSIS_MODEL", "gemma4:31b")

REQUEST_DELAY = 2  # Пауза между запросами (2 секунды)

logger.info("="*80)
logger.info("ЗАПУСК АНАЛИЗА ЧЕРТЕЖЕЙ")
logger.info("="*80)
logger.info(f"Модель: {ANALYSIS_MODEL}")
logger.info(f"🔌 Подключение к Ollama: {OLLAMA_URL}")


def get_analysis_prompt(question: str) -> str:
    """Формирует промпт анализа с обязательным указанием размеров и материалов"""
    return f"""Ты - эксперт по строительным чертежам с 20-летним опытом.

ПОЛЬЗОВАТЕЛЬ ЗАПРАШИВАЕТ: "{question}"

ЗАМЕТКА: Запрос может быть:
- Классом элемента из справочника (например: "Стена", "Колонна", "Перегородка", "Фундамент" и т.д.)
- Произвольным описанием (например: "несущие стены", "лестницы ширина и материал", "утеплитель где используется")

Найди на чертеже ВСЮ информацию, связанную с этим запросом.


❗ КРИТИЧЕСКИ ВАЖНЫЕ ТРЕБОВАНИЯ (НЕ ИГНОРИРУЙ!):


Для КАЖДОГО найденного объекта ТЫ ОБЯЗАН указать:

1. ЛИНЕЙНЫЙ РАЗМЕР (поле "dimensions"):
   - Для стен: толщина в мм (например: "толщина 380 мм")
   - Для колонн: сечение в мм (например: "400x400 мм")
   - Для балок: высота и ширина (например: "высота 450 мм, ширина 200 мм")
   - Для дверей/окон: ширина проема (например: "ширина 900 мм")
   - Для лестниц: ширина марша (например: "ширина 1200 мм")
   - Для перекрытий: толщина (например: "толщина 220 мм")
   - Для фундаментов: ширина подошвы (например: "ширина 600 мм")
   
   ЕСЛИ РАЗМЕР НЕ УКАЗАН НА ЧЕРТЕЖЕ - НАПИШИ "размер не указан на чертеже"

2. МАТЕРИАЛ (поле "material"):
   - Из чего сделан объект (кирпич, бетон, железобетон, металл, дерево, газоблок и т.д.)
   - Если есть марка материала - укажи (М150, В25, С245, D500 и т.д.)
   
   ЕСЛИ МАТЕРИАЛ НЕ УКАЗАН НА ЧЕРТЕЖЕ - НАПИШИ "материал не указан на чертеже"

================================================================================

ТРЕБОВАНИЯ К ПОИСКУ:
1. Внимательно посмотри на чертеж
2. Найди ВСЕ объекты, соответствующие запросу "{question}"
   - Если запрос - класс элемента (Стена, Колонна, Фундамент и т.д.), ищи все объекты этого типа
   - Если запрос содержит дополнительные условия (например "ширина и материал"), обязательно укажи эти параметры
3. Используй легенду и таблицу условных обозначений, если они есть
4. Связывай графические обозначения (штриховки) с текстовыми подписями

ВЕРНИ ОТВЕТ ТОЛЬКО В ФОРМАТЕ JSON (БЕЗ ЛИШНЕГО ТЕКСТА):

{{
  "user_query": "{question}",
  
  "analysis_info": {{
    "timestamp": "время анализа",
    "pdf_file": "имя файла"
  }},
  
  "found_objects": [
    {{
      "object_id": "1",
      "name": "название объекта (стена/колонна/дверь/лестница и т.д.)",
      "characteristics": {{
        "dimensions": "ОБЯЗАТЕЛЬНО: линейные размеры в мм или м (если нет - напиши 'размер не указан на чертеже')",
        "material": "ОБЯЗАТЕЛЬНО: материал изготовления (если нет - напиши 'материал не указан на чертеже')",
        "location": "расположение на чертеже (оси, координаты, этаж)",
        "quantity": "количество одинаковых элементов",
        "additional_params": "дополнительные параметры (ГОСТ, серия, класс бетона и т.д.)"
      }},
      "confidence": "высокая/средняя/низкая - насколько уверен в ответе"
    }}
  ],
  
  "not_found": "Если ничего не найдено по запросу - напиши 'По запросу \"{question}\" ничего не найдено на чертеже'",
  
  "detailed_answer": "МАКСИМАЛЬНО ПОДРОБНЫЙ ОТВЕТ на русском языке. Для каждого найденного объекта опиши его линейные размеры и материал. Если данные отсутствуют на чертеже - честно укажи это.",
  
  "summary": {{
    "total_found": "количество найденных объектов",
    "key_insights": ["ключевые выводы по запросу (с указанием размеров и материалов)"]
  }}
}}

НЕ ЗАБУДЬ: КАЖДЫЙ ОБЪЕКТ ДОЛЖЕН ИМЕТЬ "dimensions" И "material"!
ОТВЕТЬ ТОЛЬКО JSON, без пояснений до или после."""


def pdf_to_base64_enhanced(pdf_path: str, dpi: int = 300) -> str:
    """
    Конвертирует PDF страницу в base64 PNG с улучшением качества для отправки модели.
    Использует PIL для улучшения контраста и резкости.
    """
    try:
        from pdf2image import convert_from_path
        
        images = convert_from_path(pdf_path, dpi=dpi, first_page=1, last_page=1)
        
        if not images:
            raise Exception("Не удалось конвертировать PDF")
        
        img = images[0]
        logger.info(f"      📐 Размер изображения: {img.size}")
        
        # Конвертируем в RGB если нужно
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # Улучшаем контраст и резкость
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.5)
        img = img.filter(ImageFilter.SHARPEN)
        
        buffer = io.BytesIO()
        img.save(buffer, format="PNG", optimize=True)
        
        size_mb = len(buffer.getvalue()) / (1024 * 1024)
        logger.info(f"      📐 Размер PNG: {size_mb:.1f} МБ")
        
        return base64.b64encode(buffer.getvalue()).decode("utf-8")
    except Exception as e:
        logger.error(f"❌ Ошибка конвертации изображения {pdf_path}: {e}")
        # Fallback к простому методу через fitz
        try:
            doc = fitz.open(pdf_path)
            page = doc[0]
            zoom = 2.0
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            doc.close()
            return base64.b64encode(img_data).decode("utf-8")
        except Exception as e2:
            logger.error(f"❌ Fallback тоже не удался: {e2}")
            return None


def save_analysis_to_excel(data: dict, question: str, page_num: int, output_folder: str) -> str:
    """Сохраняет результаты анализа в Excel файл с размерами и материалами"""
    
    safe_query = question.replace(' ', '_').replace(',', '').replace('"', '').replace('?', '').replace('\\', '')[:40]
    excel_path = os.path.join(output_folder, f"page_{page_num}_{safe_query}.xlsx")
    
    # Стили Excel
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Лист 1: Найденные объекты
        if "found_objects" in data and data["found_objects"]:
            objects_data = []
            for obj in data["found_objects"]:
                chars = obj.get('characteristics', {})
                objects_data.append({
                    "ID": obj.get('object_id', ''),
                    "Объект": obj.get('name', ''),
                    "Линейный размер": chars.get('dimensions', 'не указан'),
                    "Материал": chars.get('material', 'не указан'),
                    "Расположение": chars.get('location', 'не указано'),
                    "Количество": chars.get('quantity', '1'),
                    "Доп. параметры": chars.get('additional_params', ''),
                    "Уверенность": obj.get('confidence', '')
                })
            
            df_objects = pd.DataFrame(objects_data)
            df_objects.to_excel(writer, sheet_name="Объекты (размеры+материалы)", index=False)
            
            worksheet = writer.sheets["Объекты (размеры+материалы)"]
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 40)
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
        
        # Лист 2: Сводка
        summary_data = []
        if "analysis_info" in data:
            for key, val in data["analysis_info"].items():
                summary_data.append({"Параметр": key, "Значение": val})
        
        summary_data.append({"Параметр": "Запрос пользователя", "Значение": question})
        summary_data.append({"Параметр": "Номер страницы", "Значение": page_num})
        summary_data.append({"Параметр": "Дата анализа", "Значение": datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
        
        if "summary" in data:
            summ = data["summary"]
            summary_data.append({"Параметр": "Всего найдено объектов", "Значение": summ.get('total_found', '0')})
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Сводка", index=False)
        
        worksheet = writer.sheets["Сводка"]
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 50
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Лист 3: Детальное описание
        if "detailed_answer" in data:
            desc_df = pd.DataFrame({"Детальное описание": [data["detailed_answer"]]})
            desc_df.to_excel(writer, sheet_name="Детальное описание", index=False)
            
            worksheet = writer.sheets["Детальное описание"]
            worksheet.column_dimensions['A'].width = 100
            for row in worksheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.row_dimensions[2].height = max(200, len(data["detailed_answer"]) // 5)
            
            worksheet['A1'].fill = header_fill
            worksheet['A1'].font = header_font
        
        # Лист 4: Статистика по материалам
        if "found_objects" in data and data["found_objects"]:
            materials_stats = {}
            for obj in data["found_objects"]:
                chars = obj.get('characteristics', {})
                mat = chars.get('material', 'не указан')
                if mat not in materials_stats:
                    materials_stats[mat] = 0
                materials_stats[mat] += 1
            
            stats_data = [{"Материал": mat, "Количество объектов": count} for mat, count in materials_stats.items()]
            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name="Статистика по материалам", index=False)
            
            worksheet = writer.sheets["Статистика по материалам"]
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 20
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
        
        # Лист 5: Ключевые выводы
        if "summary" in data and data["summary"].get('key_insights'):
            insights_data = [{"Вывод": insight} for insight in data["summary"]['key_insights']]
            df_insights = pd.DataFrame(insights_data)
            df_insights.to_excel(writer, sheet_name="Ключевые выводы", index=False)
            
            worksheet = writer.sheets["Ключевые выводы"]
            worksheet.column_dimensions['A'].width = 80
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
    
    return excel_path


def save_analysis_to_csv(data: dict, question: str, page_num: int, output_folder: str) -> str:
    """Сохраняет результаты анализа в CSV файл"""
    
    safe_query = question.replace(' ', '_').replace(',', '').replace('"', '').replace('?', '').replace('\\', '')[:40]
    csv_path = os.path.join(output_folder, f"page_{page_num}_{safe_query}.csv")
    
    if "found_objects" in data and data["found_objects"]:
        with open(csv_path, 'w', encoding='utf-8-sig') as f:
            f.write("ID;Объект;Линейный размер;Материал;Расположение;Количество;Доп.параметры;Уверенность\n")
            for obj in data["found_objects"]:
                chars = obj.get('characteristics', {})
                f.write(f"{obj.get('object_id', '')};"
                       f"{obj.get('name', '')};"
                       f"{chars.get('dimensions', '')};"
                       f"{chars.get('material', '')};"
                       f"{chars.get('location', '')};"
                       f"{chars.get('quantity', '')};"
                       f"{chars.get('additional_params', '')};"
                       f"{obj.get('confidence', '')}\n")
    
    return csv_path


def save_analysis_to_txt(data: dict, question: str, page_num: int, output_folder: str) -> str:
    """Сохраняет результаты анализа в текстовый файл"""
    
    safe_query = question.replace(' ', '_').replace(',', '').replace('"', '').replace('?', '').replace('\\', '')[:40]
    txt_path = os.path.join(output_folder, f"page_{page_num}_{safe_query}.txt")
    
    report_lines = []
    report_lines.append("="*100)
    report_lines.append(f"РЕЗУЛЬТАТ ПО ЗАПРОСУ: {question}")
    report_lines.append("="*100)
    report_lines.append(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append(f"Страница: {page_num}")
    report_lines.append("="*100)
    report_lines.append("")
    report_lines.append("Для каждого объекта указаны ЛИНЕЙНЫЕ РАЗМЕРЫ и МАТЕРИАЛ")
    report_lines.append("")
    
    if "error" in data:
        report_lines.append("⚠️ ОШИБКА ПАРСИНГА")
        report_lines.append("-"*50)
        report_lines.append(data.get("raw_response", "Неизвестная ошибка")[:2000])
        return "\n".join(report_lines)
    
    if "detailed_answer" in data:
        report_lines.append("ПОДРОБНЫЙ ОТВЕТ:")
        report_lines.append("-"*100)
        report_lines.append(data["detailed_answer"])
        report_lines.append("")
        report_lines.append("")
    
    if "found_objects" in data and data["found_objects"]:
        report_lines.append("НАЙДЕННЫЕ ОБЪЕКТЫ (с размерами и материалами):")
        report_lines.append("-"*100)
        
        for obj in data["found_objects"]:
            report_lines.append(f"\n   [{obj.get('object_id', '?')}] {obj.get('name', 'Объект')}")
            report_lines.append(f"   {'─'*70}")
            
            chars = obj.get('characteristics', {})
            
            dims = chars.get('dimensions', 'не указаны')
            report_lines.append(f"   ЛИНЕЙНЫЙ РАЗМЕР: {dims}")
            
            mat = chars.get('material', 'не указан')
            report_lines.append(f"   МАТЕРИАЛ: {mat}")
            
            if chars.get('location'):
                report_lines.append(f"   Расположение: {chars['location']}")
            if chars.get('quantity'):
                report_lines.append(f"   Количество: {chars['quantity']}")
            if chars.get('additional_params'):
                report_lines.append(f"   Доп. параметры: {chars['additional_params']}")
            
            report_lines.append(f"   Уверенность: {obj.get('confidence', 'не указана')}")
        
        report_lines.append("")
    
    if data.get("not_found") and data["not_found"] != "Не найдено на чертеже":
        report_lines.append(f"{data['not_found']}")
        report_lines.append("")
    
    if "summary" in data:
        summ = data["summary"]
        report_lines.append("📊 КЛЮЧЕВЫЕ ВЫВОДЫ:")
        report_lines.append("-"*100)
        report_lines.append(f"   • Всего найдено: {summ.get('total_found', 'не указано')}")
        
        if summ.get('key_insights'):
            for insight in summ['key_insights']:
                report_lines.append(f"   • {insight}")
    
    report_lines.append("")
    report_lines.append("="*100)
    report_lines.append("Размеры и материалы указаны для каждого найденного объекта")
    report_lines.append("="*100)
    
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(report_lines))
    
    return txt_path


def parse_json_response(result_text: str) -> dict:
    """Парсит JSON ответ от модели с очисткой от лишнего текста"""
    
    # Очищаем JSON от markdown обёрток
    if '```json' in result_text:
        result_text = result_text.split('```json')[1].split('```')[0]
    elif '```' in result_text:
        result_text = result_text.split('```')[1].split('```')[0]
    
    result_text = result_text.strip()
    
    try:
        data = json.loads(result_text)
        return data
    except json.JSONDecodeError as e:
        # Пытаемся найти JSON в конце текста
        import re
        json_match = re.search(r'\{.*\}(?=\s*$)', result_text, re.DOTALL)
        if json_match:
            try:
                data = json.loads(json_match.group())
                return data
            except:
                pass
        
        # Возвращаем ошибку парсинга
        return {
            "error": "JSON parsing failed",
            "raw_response": result_text[:5000],
            "detailed_answer": result_text
        }


def analyze_pages(images: List[Dict[str, Any]], question: str, output_folder: str) -> List[Dict[str, Any]]:
    """
    Анализирует страницы с чертежами для ответа на вопрос пользователя.
    СОХРАНЕНИЕ В EXCEL, JSON, TXT, CSV С ОБЯЗАТЕЛЬНЫМ УКАЗАНИЕМ ЛИНЕЙНЫХ РАЗМЕРОВ И МАТЕРИАЛОВ
    
    Args:
        images: Список словарей с информацией об изображениях (path, page_num, source_file, size)
        question: Вопрос пользователя
        output_folder: Папка для сохранения результатов анализа
        
    Returns:
        Список результатов анализа для каждого изображения
    """
    # Используем output_folder напрямую (без подпапки analysis)
    results_folder = Path(output_folder)
    results_folder.mkdir(parents=True, exist_ok=True)
    
    client = ollama.Client(host=OLLAMA_URL, timeout=300.0)
    
    results = []
    total_images = len(images)
    
    logger.info(f"🔍 Анализ {total_images} страниц для вопроса: {question[:50]}...")
    logger.info(f"💾 Результаты будут сохранены в Excel, JSON, TXT, CSV")
    
    for i, img_info in enumerate(images):
        try:
            image_path = img_info.get('path') or img_info.get('image_path')
            page_num = img_info.get('page_num', 1)
            source_file = img_info.get('source_file', 'unknown')
            size = img_info.get('size', 'unknown')
            
            logger.info(f"   Обработка страницы {i+1}/{total_images}: {os.path.basename(image_path)}")
            
            # Формируем промпт для анализа
            analysis_prompt = get_analysis_prompt(question)
            
            # Конвертируем PDF в base64 с улучшением качества
            logger.info(f"      🧠 Конвертация PDF страницы в base64 (enhanced)...")
            b64 = pdf_to_base64_enhanced(image_path)
            if not b64:
                raise Exception("Не удалось конвертировать PDF в изображение")
            
            # Отправляем запрос к модели
            logger.info(f"      🧠 Отправка запроса модели {ANALYSIS_MODEL}...")
            start_time = time.time()
            
            response = client.chat(
                model=ANALYSIS_MODEL,
                messages=[{
                    'role': 'user',
                    'content': analysis_prompt,
                    'images': [b64]
                }],
                stream=False,
                options={
                    'temperature': 0.0,
                    'num_predict': 12000,
                    'top_k': 10,
                    'top_p': 0.9,
                }
            )
            
            elapsed_time = time.time() - start_time
            logger.info(f"      ⏱️ Модель отвечала {elapsed_time:.1f} секунд")
            
            result_text = response['message']['content'].strip()
            
            # Парсим JSON ответ
            data = parse_json_response(result_text)
            
            # Добавляем мета-информацию
            data['page_num'] = page_num
            data['source_file'] = source_file
            
            # Сохраняем результаты во всех форматах
            logger.info(f"      💾 Сохранение результатов...")
            
            # 1. JSON
            json_filename = f"analysis_page_{page_num}_{source_file.replace('.pdf', '')}.json"
            json_path = results_folder / json_filename
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"         ✅ JSON: {json_path}")
            
            # 2. Excel
            try:
                excel_path = save_analysis_to_excel(data, question, page_num, str(results_folder))
                logger.info(f"         ✅ Excel: {excel_path}")
            except Exception as e:
                logger.error(f"         ⚠️ Ошибка сохранения Excel: {e}")
            
            # 3. TXT
            try:
                txt_path = save_analysis_to_txt(data, question, page_num, str(results_folder))
                logger.info(f"         ✅ TXT: {txt_path}")
            except Exception as e:
                logger.error(f"         ⚠️ Ошибка сохранения TXT: {e}")
            
            # 4. CSV
            try:
                csv_path = save_analysis_to_csv(data, question, page_num, str(results_folder))
                logger.info(f"         ✅ CSV: {csv_path}")
            except Exception as e:
                logger.error(f"         ⚠️ Ошибка сохранения CSV: {e}")
            
            results.append({
                'page_num': page_num,
                'source_file': source_file,
                'size': size,
                'image_path': image_path,
                'analysis': data.get('detailed_answer', result_text),
                'found_objects': data.get('found_objects', []),
                'result_file': str(json_path),
                'excel_file': excel_path if 'excel_path' in locals() else None,
                'txt_file': txt_path if 'txt_path' in locals() else None,
                'csv_file': csv_path if 'csv_path' in locals() else None,
                'relevant': True
            })
            
            logger.info(f"      ✅ Проанализировано и сохранено: страница {page_num} из {source_file}")
            
            # Пауза между запросами для стабильности
            if i < total_images - 1:
                time.sleep(REQUEST_DELAY)
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа изображения {img_info.get('path', img_info.get('image_path', 'unknown'))}: {e}")
            import traceback
            traceback.print_exc()
            results.append({
                'page_num': img_info.get('page_num', 1),
                'source_file': img_info.get('source_file', 'unknown'),
                'size': img_info.get('size', 'unknown'),
                'image_path': img_info.get('path', img_info.get('image_path', '')),
                'analysis': f"Ошибка анализа: {str(e)}",
                'result_file': None,
                'excel_file': None,
                'txt_file': None,
                'csv_file': None,
                'relevant': False,
                'error': str(e)
            })
    
    logger.info("\n" + "="*80)
    logger.info(f"АНАЛИЗ ЗАВЕРШЕН: {len(results)} страниц обработано")
    logger.info(f"Все результаты сохранены в папке: {results_folder}")
    logger.info("="*80)
    
    return results
